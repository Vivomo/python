import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl

fileName = r'E:\git\pythonCode\src\test.xls'
fileName2 = r'E:\git\pythonCode\src\ignore\data.xlsx'
bk = xlrd.open_workbook(fileName2)
remarks = []
lines = []
remark = ''

try:
    sh = bk.sheet_by_name("家庭亲子跑选手信息")
    # 获取行数
    nrows = sh.nrows
    # 获取列数
    ncols = sh.ncols
    print("nrows %d, ncols %d" % (nrows, ncols))

    # 获取各行数据

    for i in range(1, nrows):
        row_data = sh.row_values(i)
        number = row_data[0]
        name = row_data[1]
        if number:
            remarks.append(remark)
            lines.append(i)
            remark = name
        else:
            remark += name

        # cell_value = sh.cell_value(i, 1)
        # xlrd.xldate_as_tuple(cell_value, 0)
        # print(xlrd.xldate_as_tuple(cell_value, 0))
        # # print(cell_value)
        # print(row_data)
    remarks = remarks[1:]
    print(remarks)

except Exception as e:
    print("no sheet in %s named 参赛者信息" % fileName)


try:
    wk = openpyxl.load_workbook(fileName2)

    ws = wk.get_sheet_by_name('家庭亲子跑选手信息')

    for index, name in enumerate(remarks):
        point = 'H%d' % (lines[index] + 1)
        print(point, name)
        ws[point] = name
        cell = ws[point]
        print(cell)

    wk.save(r'E:\git\pythonCode\src\ignore\data2.xlsx')
except Exception as e:
    print(e)
